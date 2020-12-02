# coding=utf-8 #

# openpyxl version 1.8.5
# 安装命令 pip install openpyxl==1.8.5

from openpyxl import load_workbook
import codecs
import sys
import os
import shutil
import datetime
import re
import traceback

# 是否支持shell命令，用于删除文件到回收站
# 如果不支持则直接删除文件
# 安装命令 pip install pypiwin32
SUPPORT_SHELL = True
try:
    from win32com.shell import shell,shellcon
except Exception as e:
    SUPPORT_SHELL = False

reload(sys)
sys.setdefaultencoding('utf8')

row_separator = '\n'
output_encoding = 'utf-8'

zhmodel = re.compile(u'[\u4e00-\u9fa5]')

class VarType():
    INT     = 1
    FLOAT   = 2
    STRING  = 3
    BOOL    = 4
    TABLE   = 5
    NULL    = 6

VarTypeMap = {
    'i' : VarType.INT,
    'f' : VarType.FLOAT,
    's' : VarType.STRING,
    'b' : VarType.BOOL,
    't' : VarType.TABLE,
    }

def log(invalue):
    try:
        print(invalue.encode('gbk',errors='ignore'))
    except Exception as e:
        try:
            invalue = invalue.decode('GB2312',errors='ignore')
            print(invalue.encode('gbk',errors='ignore'))
        except Exception as e:
            print('--------------------------------log error------------------------------')
            print(invalue)


if not SUPPORT_SHELL:
    log('不支持删除到回收站功能,如果需要请输入 \'pip install pypiwin32\' 安装相应模块')

def delDir(dirname):
    if SUPPORT_SHELL:
        shell.SHFileOperation((0,shellcon.FO_DELETE,dirname,None, shellcon.FOF_SILENT | shellcon.FOF_ALLOWUNDO | shellcon.FOF_NOCONFIRMATION,None,None))  #删除文件到回收站
    else:
        shutil.rmtree(dirname)

def getType(value):
    if value == None:
        return VarType.NULL
    for var in VarTypeMap:
        if value.startswith(var):
            return VarTypeMap[var]
    return VarType.NULL

def getTypeMark(value):
    for var in VarTypeMap:
        if value.startswith(var):
            return var

def clearTypeMark(value):
    typeMark = getTypeMark(value)
    if typeMark == None:
        return value
    return value[len(typeMark):]


def encodeToUtf8(invalue):
    try:
        return invalue.decode('GB2312',errors='ignore')
    except Exception as e:
        log('字符串\'{}\' 转换为utf-8失败'.format(invalue))
        return invalue

# fileName : 文件名
# ignoreSheetTag : 忽略页名标记
#  rowStartTag : 有效行开始标记
def parseXLSX(fileName, ignoreSheetTag, rowStartTag):
    workbook = load_workbook(fileName, True, False, True, True)
    sheetInfoArr = []
    sheet_names = workbook.get_sheet_names()
    fileName = encodeToUtf8(fileName)
    for sheet_name in sheet_names:
        # 忽略某些子页
        if isinstance(ignoreSheetTag, str) and sheet_name.startswith(ignoreSheetTag):
            continue
        # 忽略包含中文文字的子页
        try:
            if zhmodel.search(sheet_name):
                # 包含中文，忽略
                log('忽略\'{}\'中文子页\'{}\''.format(fileName, sheet_name))
                continue
            else:
                pass
        except Exception as e:
            pass

        sheet = workbook.get_sheet_by_name(sheet_name)
        
        # key: (int)行下标  value:(bool)为True时代码该行为空
        emptyTagArr = []
        # 行信息
        rowInfoArr = []
        # 有效信息开始行
        count_key = -1
        # 行数据数组
        rowDataArr = []
        count_row = 0
        
        for row in sheet.iter_rows():
            count_col = 0
            for cell in row:
                if count_col == 0:
                    emptyTagArr.append(cell.internal_value == None)
                count_col = count_col + 1
    
        count_row = -1
        # 遍历行
        for row in sheet.iter_rows():
            count_row = count_row + 1
            # 跳过空行
            if emptyTagArr[count_row]:
                continue
            # 遍历列
            curRowData = []
            columnIndex = -1
            for column in row:
                columnIndex = columnIndex + 1
                column_value = column.internal_value
                # 找到开始行
                if count_key < 0 and isinstance(column_value, unicode) and column_value.startswith(rowStartTag):
                    column_value = column_value.strip()
                    column_value = column_value.strip(rowStartTag)
                    stepCount = 2
                    if column_value.isdigit() and int(column_value) > 0:
                        stepCount = int(column_value)
                    count_key = count_row + stepCount    
                if count_key > 0 and count_row == count_key:
                    rowInfoArr.append({'name' : column_value, 'type' : getType(column_value)})
                if count_key > 0 and count_row > count_key:
                    curRowData.append(column_value)
            if count_key > 0 and count_row > count_key:
                rowDataArr.append(curRowData)
        
        if count_key < 0:
            log('\n\n')
            log('错误:文件[\'{}\']的工作表[\'{}\']中找不到开始标记\'{}\''.format(fileName, sheet_name, rowStartTag))
            log('\n\n')

        sheetInfoArr.append({
            'sheetName' : sheet_name,
            'rowInfoArr' : rowInfoArr,
            'rowDataArr' : rowDataArr,
        })
    return sheetInfoArr



def fmtLuaValue(value, valueType):
    if valueType == VarType.INT:
        if value == None:
            return 0
        return str(int(value))
    elif valueType == VarType.FLOAT:
        if value == None:
            return 0.0
        return str(float(value))
    elif valueType == VarType.STRING:
        if value == None:
            return '\'\''
        value = str(value)
        value = checkUnescaped(value, '\'')
        value = value.replace('\n', '\\n')
        return '\'{}\''.format(value)
    elif valueType == VarType.BOOL:
        if value == None:
            return 'false'
        value = str(value)
        value = value.strip()
        value = value.lower()
        # false 0 或空都为false 其余都为true
        if value == 'false' or value == '0' or value == '':
            return 'false'
        return 'true'
    elif valueType == VarType.TABLE:
        if value == None:
            return '{}'
    return str(value)

def fmtLuaKey(value, ignoreTypeMark, strType):
    if value.isdigit():
        return value
    if ignoreTypeMark:
        if isinstance(value, unicode) or isinstance(value,str):
            value = clearTypeMark(value)
    if strType:
        return '\'{}\''.format(value)
    return '{}'.format(value)

# 判断数组是否有重复的元素并返回重复元素集
def judgeRepeated(array):
    nums={}
    repeatArr = []
    for i in array:
        if i not in nums:
            nums[i] = True
        else:
            repeatArr.append(i)
    return repeatArr

# 检查转义符
def checkUnescaped(strValue, character):
    begin = 0
    pos = 0
    while True:
        pos = strValue.find(character, begin)
        if pos < 0:
            break
        if pos == 0 or strValue[pos - 1:pos] != "\\":
            strValue = strValue[0:pos] + "\\" + strValue[pos:]
            begin = pos + 2
        else:
            begin = pos + 1
    return strValue

def getBaseName(filePath):
    filePath = os.path.split(filePath)[1]
    return os.path.splitext(filePath)[0]

# 导出
def exportToLua(luaFile, xlsxFileName, rootDir, ignoreFirstColumn, ignoreTypeMark, outFileArr):
    sheetInfoArr = parseXLSX(xlsxFileName, '_', '#begin')

    #######################################################################################
    #if 以xlsxFileName命名导出lua文件
    #######################################################################################

    # luaFileBaseName = getBaseName(luaFile)
    # xlsxFileBaseName = getBaseName(xlsxFileName)
    # outLuaFile = luaFile
    # isFirstSheet = True
    # for sheetInfo in sheetInfoArr:
    #     log(sheetInfo['sheetName'] + ' start...')

    #     if len(sheetInfoArr) == 1:
    #         pass
    #     else:
    #         if isFirstSheet and sheetInfo['sheetName'] == luaFileBaseName:
    #             isFirstSheet = False
    #             outLuaFile = luaFile
    #         else:
    #             outLuaFile = os.path.split(luaFile)[0] + '/' + luaFileBaseName + '_' + sheetInfo['sheetName'] + os.path.splitext(luaFile)[1]
    #     outFileArr.append(outLuaFile)

    #######################################################################################
    #else
    #######################################################################################

    for sheetInfo in sheetInfoArr:
        log(sheetInfo['sheetName'] + ' start...')

        outLuaFile = os.path.split(luaFile)[0] + '/' + sheetInfo['sheetName'] + os.path.splitext(luaFile)[1]
        outFileArr.append(outLuaFile)

    #######################################################################################
    #endif
    #######################################################################################

        rowInfo = sheetInfo['rowInfoArr']
        colCount = len(rowInfo)
        tabKeysArr = []
        scriptLineArr = []
        for rowData in sheetInfo['rowDataArr']:
            rowScript = ''
            isFirstEle = True
            for idx in xrange(0, colCount):
                rowValue = None
                try:
                    rowValue = rowData[idx]
                except Exception as e:
                    rowValue = None
                curName = rowInfo[idx]['name']
                curType = rowInfo[idx]['type']
                # 
                tmpValue = str(fmtLuaValue(rowValue, curType))
                if curType != VarType.NULL:
                    if isFirstEle:
                        isFirstEle = False
                        rowScript = rowScript + 'M.Data[{}] = {{'.format(fmtLuaKey(tmpValue, ignoreTypeMark, False))
                        tabKeysArr.append(fmtLuaKey(tmpValue, ignoreTypeMark, False))
                        if not ignoreFirstColumn:
                            rowScript = rowScript + '{} = {}, '.format(fmtLuaKey(curName, ignoreTypeMark, False), tmpValue)
                    else:
                        rowScript = rowScript + '{} = {}, '.format(fmtLuaKey(curName, ignoreTypeMark, False), tmpValue)
            scriptLineArr.append(rowScript.rstrip(', ') + '}' + row_separator)

        repeatArr = judgeRepeated(tabKeysArr)
        excelFileName = encodeToUtf8(xlsxFileName.replace(rootDir, '').replace('\\', '/'))
        if len(repeatArr) > 0:
            log('\n\n')
            log('错误:文件{}导出失败'.format(outLuaFile)) 
            for repeatKey in repeatArr:
                log('错误:文件[\'{}\']的工作表[\'{}\']中含有重复Key \'{}\''.format(excelFileName, sheetInfo['sheetName'], repeatKey))
            log('\n\n')
            return False
        else:
            script = '-- Auto Generate by Excel \'{}\', Don\'t try to modify!'.format(excelFileName) + row_separator
            # script = script + '-- Generated on {}.'.format(datetime.datetime.strftime(datetime.datetime.now(),'%Y-%m-%d %H:%M:%S')) + row_separator
            script = script + row_separator

            script = script + 'local M = {}' + row_separator
            script = script + row_separator

            script = script + 'M.Data = {}' + row_separator
            lineCount = 0
            for line in scriptLineArr:
                script = script + line
                lineCount = lineCount + 1

            script = script + 'M.length = {}'.format(lineCount) + row_separator
            script = script + row_separator

            script = script + 'function M.getData(key)' + row_separator
            script = script + '    return M.Data[key]' + row_separator
            script = script + 'end' + row_separator
            script = script + row_separator

            script = script + 'function M.getItem(key1, key2)' + row_separator
            script = script + '    if M.Data[key1] == nil then' + row_separator
            script = script + '        print(string.format("错误: 文件\'{}\'中配置\'%s\'不存在!!!!!!!!!!!!!!!!!!!!!!!", tostring(key1)))'.format(os.path.basename(outLuaFile)) + row_separator
            script = script + '        return' + row_separator
            script = script + '    end' + row_separator
            script = script + '    return M.Data[key1][key2]' + row_separator
            script = script + 'end' + row_separator
            script = script + row_separator

            script = script + 'function M.getAllData()' + row_separator
            script = script + '    return M.Data' + row_separator
            script = script + 'end' + row_separator
            script = script + row_separator

            script = script + 'return M' + row_separator
            
            try:
                script = script.encode('utf-8',errors='ignore')
            except Exception as e:
                log('错误:文件{} script转换编码失败'.format(outLuaFile))
                # log('\n\n\n')
                # log(script)
                # log('\n\n\n')
            try:
                output_file = codecs.open(outLuaFile, 'w', output_encoding)
                output_file.write(script)
                output_file.close()
            except Exception as e:
                log('错误:文件{}写入失败'.format(outLuaFile))
                log('\n\n\n')
                traceback.print_exc()
                log('\n\n\n')
                return False
    return True

# 遍历文件夹
def exportDir(workDir, outDir):
    ExtensionArr = ['.xlsx', '.xlsm']
    exportExt = '.lua'

    # 删除输出文件夹
    if workDir != outDir:
        if os.path.isdir(outDir):
            log('删除文件夹 {}'.format(outDir))
            delDir(outDir)

    log("开始发布>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
    outFileArr = []
    for root, dirs, files in os.walk(workDir):
        # 遍历文件
        for f in files:
            for ext in ExtensionArr:
                if f.endswith(ext):
                    fullPath = f[:-len(ext)] + exportExt
                    fullPath = os.path.join(root, fullPath)
                    fullPath = outDir + fullPath[len(workDir):]

                    out = outDir + root[len(workDir):]
                    if not os.path.isdir(out):
                        os.makedirs(out)
                    if not exportToLua(fullPath, os.path.join(root, f), workDir, True, True, outFileArr):
                        log("发布失败<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<")
                        return False


    log("开始优化>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
    for file in outFileArr:
        fsize = os.path.getsize(os.path.abspath(file))
        # 大于100kb才进行优化
        if fsize > 1024 * 100:
            print("%s start..." % file)
            luaexe = os.path.abspath("luajit.exe")
            luafile = os.path.abspath("DataTableOptimizer.lua")
            os.system("%s %s %s %s" % (luaexe, luafile, file, os.path.abspath(file)))
    log("优化结束<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<")

    script = 'local M = {}' + row_separator

    for file in outFileArr:
        file = file.replace('.lua', '')
        baseKey = file.replace(outDir, '')
        baseKey = baseKey.replace('/', '_')
        baseKey = baseKey.replace('\\', '_')
        baseLua = file.replace(outDir, '')
        baseLua = baseLua.replace('/', '.')
        baseLua = baseLua.replace('\\', '.')
        script = script + 'M.{} = require(\'XXConfig.{}\')'.format(baseKey, baseLua) + row_separator
    script = script + 'return M' + row_separator
    
    output_file = codecs.open(outDir + 'XXConfig.lua', 'w', output_encoding)
    output_file.write(script)
    output_file.close()

    log("发布成功<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<")
    return True

def copy_path(fromPath,toPath):
    if os.path.exists(toPath):
        try:
            delDir(toPath)
            log('delete ' + toPath + ' succeeded')
        except Exception as e:
            log('delete ' + toPath + ' failed')
            return False
    if os.path.exists(fromPath):
        try:
            shutil.copytree(fromPath, toPath)
        except Exception as e:
            return False
    return True

def main(argv):
    # workDir = './sheet/'
    # tmpDir = './tmp/'
    # export = ['./../../Parkour_310/src/XXConfig/']

    if len(argv) < 4:
        log("启动参数错误")
        log("使用方法: python Excel2Lua.py 文档路径 临时缓存路径 发布路径1 发布路径2 ...(支持多路径发布)")
        return

    workDir = argv[1]
    tmpDir = argv[2]
    export = []

    for i in range(3,len(argv)):
        export.append(argv[i])

    if not exportDir(workDir, tmpDir):
        return
    log('\n\n')
    for arg in export:
        if copy_path(tmpDir, arg):
            log('copy to ' + arg + ' succeeded')
        else:
            log('copy to ' + arg + ' failed')

if __name__ == '__main__':
    try:
        main(sys.argv)
    except Exception as e:
        log('发布出错:')
        log(e)
        traceback.print_exc()
    finally:
        os.system('pause')
        pass
    
